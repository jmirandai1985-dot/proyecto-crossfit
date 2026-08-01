"""
Servicio de Reportes Excel - Version Premium 4 pestanas
Genera reportes ejecutivos con diseno profesional, graficos y formulas reales
"""
import io
from datetime import datetime, date, timedelta, timezone
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


# ============================================================
# COLORES CORPORATIVOS (diseno aprobado)
# ============================================================
NAVY = "111827"
NAVY_HEADER = "1F3A5F"
ORANGE_ACCENT = "EA580C"
BLUE_KPI = "2563EB"
GREEN_KPI = "16A34A"
PURPLE_KPI = "7C3AED"
TEAL_KPI = "0D9488"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
MEDIUM_GRAY = "D1D5DB"
DARK_TEXT = "1F2937"
RED = "DC2626"
GREEN = "16A34A"
YELLOW_BG = "FEF3C7"

# Barras de color para KPI cards
KPI_COLORS = {
    "ingresos": ORANGE_ACCENT,
    "alumnos": BLUE_KPI,
    "mrr": GREEN_KPI,
    "arpu": PURPLE_KPI,
    "nuevos": TEAL_KPI,
}

# Estilos reutilizables
header_fill = PatternFill(start_color=NAVY_HEADER,
                          end_color=NAVY_HEADER, fill_type="solid")
header_font = Font(bold=True, color=WHITE, size=11)
title_font = Font(bold=True, size=18, color=NAVY)
section_font = Font(bold=True, size=13, color=NAVY)
kpi_value_font = Font(bold=True, size=24, color=DARK_TEXT)
kpi_label_font = Font(bold=True, size=11, color=DARK_TEXT)
kpi_subtitle_font = Font(size=9, color="6B7280")
thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY),
)
alt_fill = PatternFill(start_color=LIGHT_GRAY,
                       end_color=LIGHT_GRAY, fill_type="solid")
total_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
total_font = Font(bold=True, color=WHITE, size=11)
green_font = Font(bold=True, color=GREEN, size=11)
red_font = Font(bold=True, color=RED, size=11)


def _apply_kpi_card(ws, row, col, label, value, subtitle, accent_color, emoji):
    """Dibuja una tarjeta KPI con barra de color superior."""
    cell = ws.cell(row=row, column=col, value=emoji + "  " + label)
    cell.font = kpi_label_font
    cell.alignment = Alignment(horizontal='left', vertical='top')
    # Barra de color arriba (fila anterior, misma col)
    ws.cell(row=row - 1, column=col).fill = PatternFill(
        start_color=accent_color, end_color=accent_color, fill_type="solid")
    ws.cell(row=row - 1, column=col).border = Border(
        bottom=Side(style='medium', color=accent_color))
    # Valor
    cell_v = ws.cell(row=row + 1, column=col, value=value)
    cell_v.font = kpi_value_font
    cell_v.alignment = Alignment(horizontal='left', vertical='top')
    # Subtitulo
    cell_s = ws.cell(row=row + 2, column=col, value=subtitle)
    cell_s.font = kpi_subtitle_font
    cell_s.alignment = Alignment(horizontal='left', vertical='top')


def _style_section_header(ws, row, col, title, accent_color):
    """Barra de acento a la izquierda + titulo de seccion."""
    ws.cell(row=row, column=col, value="").fill = PatternFill(
        start_color=accent_color, end_color=accent_color, fill_type="solid")
    ws.cell(row=row, column=col).border = Border(
        right=Side(style='medium', color=accent_color))
    cell_title = ws.cell(row=row, column=col + 1, value=title)
    cell_title.font = section_font
    cell_title.alignment = Alignment(horizontal='left', vertical='center')


def _style_table_header(ws, row, headers, start_col=1):
    """Encabezado de tabla con fondo navy y texto blanco."""
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)


def _style_data_row(ws, row, values, start_col=1, is_alt=False, is_percent_cols=None, is_money_cols=None):
    """Fila de datos con banding y formato."""
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if is_alt:
            cell.fill = alt_fill
        col_idx = start_col + i
        if is_percent_cols and col_idx in is_percent_cols:
            cell.number_format = '0.0%'
        if is_money_cols and col_idx in is_money_cols:
            cell.number_format = '#,##0'
    return row


def _add_bar_chart(ws, data_rows, data_col, cat_col, title, position, color=ORANGE_ACCENT):
    """Agrega grafico de barras."""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.width = 13
    chart.height = 8
    data_ref = Reference(ws, min_col=data_col,
                         min_row=data_rows[0], max_row=data_rows[1])
    cats_ref = Reference(ws, min_col=cat_col,
                         min_row=data_rows[0], max_row=data_rows[1])
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = color
    ws.add_chart(chart, position)


def _add_pie_chart(ws, data_rows, data_col, cat_col, title, position):
    """Agrega grafico de torta."""
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = 13
    chart.height = 8
    data_ref = Reference(ws, min_col=data_col,
                         min_row=data_rows[0], max_row=data_rows[1])
    cats_ref = Reference(ws, min_col=cat_col,
                         min_row=data_rows[0], max_row=data_rows[1])
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False
    ws.add_chart(chart, position)


def _add_line_chart(ws, data_rows, data_col, cat_col, title, position, color=ORANGE_ACCENT):
    """Agrega grafico de linea."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = 13
    chart.height = 8
    data_ref = Reference(ws, min_col=data_col,
                         min_row=data_rows[0], max_row=data_rows[1])
    cats_ref = Reference(ws, min_col=cat_col,
                         min_row=data_rows[0], max_row=data_rows[1])
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = color
    chart.series[0].graphicalProperties.line.solidFill = color
    ws.add_chart(chart, position)


def _build_historico_mensual(db, tenant_id):
    """
    Construye la tabla Historico Mensual con datos reales de transacciones_financieras
    para los ultimos 6 meses.
    Retorna: list of dicts con Año, Mes, MesNum, IngresosTotales, Egresos, Neto
    """
    historico = []
    ahora = datetime.now(timezone.utc)
    for i in range(5, -1, -1):
        # Calcular mes
        m = ahora.month - i
        y = ahora.year
        while m <= 0:
            m += 12
            y -= 1
        inicio = date(y, m, 1)
        if m == 12:
            fin = date(y + 1, 1, 1) - timedelta(days=1)
        else:
            fin = date(y, m + 1, 1) - timedelta(days=1)

        # Ingresos
        ing = db.execute(text("""
            SELECT COALESCE(SUM(monto), 0) FROM transacciones_financieras
            WHERE tenant_id = :tid AND tipo = 'ingreso' AND fecha >= :ini AND fecha <= :fin
        """), {"tid": tenant_id, "ini": inicio, "fin": fin}).scalar() or 0

        # Egresos
        eg = db.execute(text("""
            SELECT COALESCE(SUM(monto), 0) FROM transacciones_financieras
            WHERE tenant_id = :tid AND tipo = 'egreso' AND fecha >= :ini AND fecha <= :fin
        """), {"tid": tenant_id, "ini": inicio, "fin": fin}).scalar() or 0

        # Alumnos activos fin de mes
        alumnos = db.execute(text("""
            SELECT COUNT(DISTINCT u.id) FROM usuarios u
            JOIN suscripciones s ON u.id = s.usuario_id
            WHERE u.tenant_id = :tid AND u.rol = 'alumno' AND u.activo = true
              AND s.estado = 'activo' AND s.fecha_expiracion >= :fin
        """), {"tid": tenant_id, "fin": fin}).scalar() or 0

        # Nuevos alumnos del mes
        nuevos = db.execute(text("""
            SELECT COUNT(*) FROM usuarios
            WHERE tenant_id = :tid AND rol = 'alumno'
              AND created_at::date >= :ini AND created_at::date <= :fin
        """), {"tid": tenant_id, "ini": inicio, "fin": fin}).scalar() or 0

        # MRR (suma de precios de planes activos)
        mrr_val = db.execute(text("""
            SELECT COALESCE(SUM(p.precio_clp), 0) FROM suscripciones s
            JOIN planes p ON s.plan_id = p.id
            WHERE s.tenant_id = :tid AND s.estado = 'activo' AND s.fecha_expiracion >= :fin
        """), {"tid": tenant_id, "fin": fin}).scalar() or 0

        # Ventas bazar del mes
        ventas = db.execute(text("""
            SELECT COALESCE(SUM(total), 0) FROM pedidos
            WHERE tenant_id = :tid AND fecha_pedido >= :ini AND fecha_pedido <= :fin
              AND estado != 'cancelado'
        """), {"tid": tenant_id, "ini": inicio, "fin": fin}).scalar() or 0

        meses_es = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                    'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        label = f"{meses_es[m - 1]} {y}"

        ing_f = float(ing)
        eg_f = float(eg)

        historico.append({
            "anio": y,
            "mes": label,
            "mes_num": m,
            "ingresos": ing_f,
            "egresos": eg_f,
            "neto": ing_f - eg_f,
            "alumnos": alumnos,
            "nuevos": nuevos,
            "mrr": float(mrr_val),
            "ventas": float(ventas),
        })
    return historico


def _categorias_por_genero(db, tenant_id):
    """Retorna resumen de planes por categoria."""
    rows = db.execute(text("""
        SELECT p.id, p.nombre, p.genero, p.es_estudiante, p.precio_clp, p.creditos, p.activo,
               (SELECT COUNT(*) FROM suscripciones s WHERE s.plan_id = p.id AND s.estado = 'activo') as susc_activas
        FROM planes p
        WHERE p.tenant_id = :tid
        ORDER BY p.id
    """), {"tid": tenant_id}).fetchall()

    categorias = {
        "masculino": {"label": "Masculino", "planes": 0, "suscripciones": 0, "ingresos": 0},
        "femenino": {"label": "Femenino", "planes": 0, "suscripciones": 0, "ingresos": 0},
        "estudiante_m": {"label": "Est. Masculino", "planes": 0, "suscripciones": 0, "ingresos": 0},
        "estudiante_f": {"label": "Est. Femenino", "planes": 0, "suscripciones": 0, "ingresos": 0},
    }

    planes_list = []
    for r in rows:
        es_est = bool(r.es_estudiante) if r.es_estudiante else False
        gen = (r.genero or "").strip().lower()
        key = None
        if es_est and gen == "masculino":
            key = "estudiante_m"
        elif es_est and gen == "femenino":
            key = "estudiante_f"
        elif gen == "masculino" or gen == "unisex":
            if es_est:
                key = "estudiante_m"
            else:
                key = "masculino"
        elif gen == "femenino":
            if es_est:
                key = "estudiante_f"
            else:
                key = "femenino"

        susc = r.susc_activas or 0
        ing_plan = susc * (r.precio_clp or 0)

        if key and key in categorias:
            categorias[key]["planes"] += 1
            categorias[key]["suscripciones"] += susc
            categorias[key]["ingresos"] += ing_plan

        planes_list.append({
            "id": r.id,
            "nombre": r.nombre,
            "genero": r.genero,
            "es_estudiante": es_est,
            "precio": r.precio_clp or 0,
            "creditos": r.creditos or 0,
            "activo": r.activo,
            "suscripciones": susc,
        })

    return categorias, planes_list


def _pedidos_bazar_mes(db, tenant_id, inicio, fin):
    """Obtiene pedidos del bazar del mes."""
    try:
        rows = db.execute(text("""
            SELECT p.id, p.alumno_id, u.nombre as alumno_nombre,
                   pr.nombre as producto_nombre, p.cantidad, p.total, p.estado, p.fecha_pedido
            FROM pedidos p
            JOIN usuarios u ON p.alumno_id = u.id
            JOIN productos pr ON p.producto_id = pr.id
            WHERE p.tenant_id = :tid AND p.fecha_pedido >= :ini AND p.fecha_pedido < :fin
            ORDER BY p.fecha_pedido DESC
        """), {"tid": tenant_id, "ini": inicio, "fin": fin}).fetchall()
        return rows
    except Exception:
        return []


def crear_reporte_ventas_mensual_bytes(
    db: Session,
    tenant_id: int,
    mes: int,
    anio: int
) -> bytes:
    """
    Genera reporte Excel con 4 pestanas:
    1. Resumen Ejecutivo
    2. Detalle Planes
    3. Bazar y Servicios
    4. Flujo de Caja
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ── PERIODOS ──
    fecha_inicio = date(anio, mes, 1)
    if mes == 12:
        fecha_fin = date(anio + 1, 1, 1)
    else:
        fecha_fin = date(anio, mes + 1, 1)
    mes_ant = mes - 1 if mes > 1 else 12
    anio_ant = anio if mes > 1 else anio - 1
    inicio_ant = date(anio_ant, mes_ant, 1)
    if mes_ant == 12:
        fin_ant = date(anio_ant + 1, 1, 1)
    else:
        fin_ant = date(anio_ant, mes_ant + 1, 1)

    MESES_ES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    label_mes = f"{MESES_ES[mes - 1]} {anio}"
    label_mes_ant = f"{MESES_ES[mes_ant - 1]} {anio_ant}"

    # ── DATOS AGREGADOS ──
    # KPI: Ingresos del mes
    ing_mes = db.execute(text("""
        SELECT COALESCE(SUM(monto), 0) FROM transacciones_financieras
        WHERE tenant_id = :tid AND tipo = 'ingreso' AND fecha >= :ini AND fecha < :fin
    """), {"tid": tenant_id, "ini": fecha_inicio, "fin": fecha_fin}).scalar() or 0
    eg_mes = db.execute(text("""
        SELECT COALESCE(SUM(monto), 0) FROM transacciones_financieras
        WHERE tenant_id = :tid AND tipo = 'egreso' AND fecha >= :ini AND fecha < :fin
    """), {"tid": tenant_id, "ini": fecha_inicio, "fin": fecha_fin}).scalar() or 0
    ing_mes_f = float(ing_mes)
    eg_mes_f = float(eg_mes)
    neto_mes = ing_mes_f - eg_mes_f

    # Mes anterior
    ing_ant = db.execute(text("""
        SELECT COALESCE(SUM(monto), 0) FROM transacciones_financieras
        WHERE tenant_id = :tid AND tipo = 'ingreso' AND fecha >= :ini AND fecha < :fin
    """), {"tid": tenant_id, "ini": inicio_ant, "fin": fin_ant}).scalar() or 0
    eg_ant = db.execute(text("""
        SELECT COALESCE(SUM(monto), 0) FROM transacciones_financieras
        WHERE tenant_id = :tid AND tipo = 'egreso' AND fecha >= :ini AND fecha < :fin
    """), {"tid": tenant_id, "ini": inicio_ant, "fin": fin_ant}).scalar() or 0
    ing_ant_f = float(ing_ant)
    eg_ant_f = float(eg_ant)
    neto_ant = ing_ant_f - eg_ant_f

    # Alumnos activos
    alumnos_activos = db.execute(text("""
        SELECT COUNT(DISTINCT u.id) FROM usuarios u
        JOIN suscripciones s ON u.id = s.usuario_id
        WHERE u.tenant_id = :tid AND u.rol = 'alumno' AND u.activo = true
          AND s.estado = 'activo' AND s.fecha_expiracion >= CURRENT_DATE
    """), {"tid": tenant_id}).scalar() or 0

    # Nuevos alumnos
    nuevos_alumnos = db.execute(text("""
        SELECT COUNT(*) FROM usuarios
        WHERE tenant_id = :tid AND rol = 'alumno'
          AND created_at::date >= :ini AND created_at::date < :fin
    """), {"tid": tenant_id, "ini": fecha_inicio, "fin": fecha_fin}).scalar() or 0

    # MRR
    mrr_val = db.execute(text("""
        SELECT COALESCE(SUM(p.precio_clp), 0) FROM suscripciones s
        JOIN planes p ON s.plan_id = p.id
        WHERE s.tenant_id = :tid AND s.estado = 'activo' AND s.fecha_expiracion >= CURRENT_DATE
    """), {"tid": tenant_id}).scalar() or 0
    mrr_val_f = float(mrr_val)

    # ARPU
    arpu_val = round(neto_mes / alumnos_activos,
                     0) if alumnos_activos > 0 else 0

    # Ocupacion por disciplina
    ocupacion = db.execute(text("""
        SELECT d.nombre,
               SUM(COALESCE(c.asistentes_confirmados, 0)) as asistentes,
               SUM(COALESCE(c.cupo_maximo, 1)) as cupo_total
        FROM clases c
        JOIN disciplinas d ON c.disciplina_id = d.id
        WHERE c.tenant_id = :tid AND c.fecha >= :ini AND c.fecha < :fin
        GROUP BY d.nombre ORDER BY d.nombre
    """), {"tid": tenant_id, "ini": fecha_inicio, "fin": fecha_fin}).fetchall()

    # Egresos del mes detalle
    egresos = db.execute(text("""
        SELECT categoria, monto, descripcion, fecha
        FROM transacciones_financieras
        WHERE tenant_id = :tid AND tipo = 'egreso' AND fecha >= :ini AND fecha < :fin
        ORDER BY fecha DESC
    """), {"tid": tenant_id, "ini": fecha_inicio, "fin": fecha_fin}).fetchall()

    # Planes y categorias
    categorias, planes_list = _categorias_por_genero(db, tenant_id)

    # Pedidos bazar
    pedidos = _pedidos_bazar_mes(db, tenant_id, fecha_inicio, fecha_fin)
    ventas_bazar = sum(p.total for p in pedidos)
    pedidos_completados = sum(1 for p in pedidos if p.estado == "entregado")
    pedidos_pendientes = sum(1 for p in pedidos if p.estado == "pendiente")

    # Ventas por producto (bazar)
    ventas_producto = {}
    for p in pedidos:
        nom = p.producto_nombre or "Producto"
        ventas_producto[nom] = ventas_producto.get(nom, 0) + p.total

    # Historico mensual
    historico = _build_historico_mensual(db, tenant_id)

    # ================================================================
    # PESTANA 1: RESUMEN EJECUTIVO
    # ================================================================
    ws1 = wb.create_sheet("Resumen Ejecutivo", 0)
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 22
    ws1.column_dimensions['F'].width = 22
    ws1.column_dimensions['G'].width = 22
    ws1.column_dimensions['H'].width = 3

    # Banner titulo
    for c in range(1, 9):
        ws1.cell(row=1, column=c).fill = PatternFill(
            start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws1.merge_cells('A1:H1')
    banner_cell = ws1.cell(
        row=1, column=1, value="REPORTE EJECUTIVO - BOX CROSSFIT")
    banner_cell.font = Font(bold=True, size=20, color=WHITE)
    banner_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 45

    # Franja naranja
    for c in range(1, 9):
        ws1.cell(row=2, column=c).fill = PatternFill(
            start_color=ORANGE_ACCENT, end_color=ORANGE_ACCENT, fill_type="solid")
    ws1.row_dimensions[2].height = 6

    # Subtitulo
    ws1.merge_cells('A3:H3')
    ws1.cell(row=3, column=1,
             value=f"Periodo: {label_mes}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(size=10, color="6B7280")
    ws1.cell(row=3, column=1).alignment = Alignment(horizontal='center')

    # ── KPI CARDS ──
    kpi_start = 5
    # Fila 4 = barra de color, Fila 5 = label, Fila 6 = valor, Fila 7 = subtitulo
    # Barras de color (fila 4)
    for c in range(2, 7):
        ws1.cell(row=kpi_start - 1, column=c).border = Border(bottom=Side(style='medium', color=KPI_COLORS.get(
            ["ingresos", "alumnos", "mrr", "arpu", "nuevos"][c - 2], ORANGE_ACCENT)))
    _apply_kpi_card(ws1, kpi_start, 2, "Ingresos del Mes",
                    f"${neto_mes:,.0f}", "Ingresos - Egresos", ORANGE_ACCENT, "💰")
    _apply_kpi_card(ws1, kpi_start, 3, "Alumnos Activos",
                    f"{alumnos_activos}", "Con suscripcion vigente", BLUE_KPI, "👥")
    _apply_kpi_card(ws1, kpi_start, 4, "MRR",
                    f"${mrr_val_f:,.0f}", "Ingresos recurrentes", GREEN_KPI, "📈")
    _apply_kpi_card(ws1, kpi_start, 5, "ARPU",
                    f"${arpu_val:,.0f}", "Por alumno activo", PURPLE_KPI, "🎯")
    _apply_kpi_card(ws1, kpi_start, 6, "Nuevos Alumnos",
                    f"{nuevos_alumnos}", "Este mes", TEAL_KPI, "✨")
    ws1.row_dimensions[kpi_start].height = 20
    ws1.row_dimensions[kpi_start + 1].height = 35
    ws1.row_dimensions[kpi_start + 2].height = 20

    # ── SECCION: Ingresos por Unidad de Negocio ──
    r = 10
    _style_section_header(
        ws1, r, 1, "Ingresos por Unidad de Negocio", BLUE_KPI)
    r += 1

    # Datos de ingresos
    # Membresias = ing_mes_f, Bazar = ventas_bazar, Arriendos = 0, Eventos = 0
    total_negocio = ing_mes_f + ventas_bazar
    negocio_items = [
        ("Membresias / Planes", ing_mes_f),
        ("Bazar / Tienda", ventas_bazar),
        ("Arriendos", 0),
        ("Eventos", 0),
    ]
    tbl_headers = ["Unidad de Negocio", "Monto",
                   "% del Total", "Var. vs Mes Ant"]
    _style_table_header(ws1, r, tbl_headers)
    r += 1
    negocio_data_start = r  # track first data row for charts
    ing_ant_total = ing_ant_f + float(db.execute(text("""
        SELECT COALESCE(SUM(total), 0) FROM pedidos
        WHERE tenant_id = :tid AND fecha_pedido >= :ini AND fecha_pedido < :fin AND estado != 'cancelado'
    """), {"tid": tenant_id, "ini": inicio_ant, "fin": fin_ant}).scalar() or 0)

    for i, (nom, monto_item) in enumerate(negocio_items):
        is_alt = i % 2 == 1
        pct = (monto_item / total_negocio * 100) if total_negocio > 0 else 0
        var = monto_item - ing_ant_total
        var_label = f"+${var:,.0f}" if var >= 0 else f"-${abs(var):,.0f}"
        _style_data_row(ws1, r, [
                        nom, monto_item, f"{pct:.1f}%", var_label], start_col=1, is_alt=is_alt, is_money_cols={2})
        r += 1

    # Total
    _style_data_row(ws1, r, ["TOTAL", total_negocio,
                    "100.0%", ""], start_col=1, is_money_cols={2})
    for c in range(1, 5):
        ws1.cell(row=r, column=c).fill = total_fill
        ws1.cell(row=r, column=c).font = total_font
    r += 2

    # Grafico de barras: ingresos por unidad
    data_start = negocio_data_start
    if data_start < 1:
        data_start = 1
    bar_end = data_start + len(negocio_items) - 1
    if data_start <= bar_end:
        _add_bar_chart(ws1, (data_start, bar_end), 2, 1,
                       "Ingresos por Unidad de Negocio", f"B{r}", BLUE_KPI)
    if data_start <= bar_end:
        _add_pie_chart(ws1, (data_start, bar_end), 2, 1,
                       "Distribucion de Ingresos", f"F{r}")
    r += 16

    # ── SECCION: Ocupacion por Disciplina ──
    _style_section_header(ws1, r, 1, "Ocupacion por Disciplina", TEAL_KPI)
    r += 1
    _style_table_header(
        ws1, r, ["Disciplina", "Asistentes", "Cupo Total", "Ocupacion"])
    r += 1
    occ_start = r
    for i, row_d in enumerate(ocupacion):
        is_alt = i % 2 == 1
        cupo = row_d.cupo_total or 0
        pct_occ = round(row_d.asistentes / cupo * 100, 1) if cupo > 0 else 0
        _style_data_row(ws1, r, [row_d.nombre, row_d.asistentes,
                        cupo, f"{pct_occ}%"], start_col=1, is_alt=is_alt)
        r += 1
    occ_end = r - 1
    r += 1

    # Grafico dona (pie chart) - skip if no data
    if occ_start <= occ_end:
        _add_pie_chart(ws1, (occ_start, occ_end), 2, 1,
                       "Ocupacion por Disciplina", f"B{r}")
    r += 14

    # ── SECCION: Tendencia Ingresos 6 Meses ──
    _style_section_header(
        ws1, r, 1, "Tendencia de Ingresos (Ultimos 6 Meses)", GREEN_KPI)
    r += 1
    _style_table_header(ws1, r, [
                        "Mes", "Ingresos", "Egresos", "Neto", "Alumnos", "Nuevos", "MRR", "Ventas"])
    r += 1
    hist_start = r
    for i, h in enumerate(historico):
        is_alt = i % 2 == 1
        _style_data_row(ws1, r, [h["mes"], h["ingresos"], h["egresos"], h["neto"], h["alumnos"], h["nuevos"], h["mrr"], h["ventas"]],
                        start_col=1, is_alt=is_alt, is_money_cols={2, 3, 4, 7, 8})
        r += 1
    hist_end = r - 1
    r += 1

    # Grafico linea: tendencia neto
    _add_line_chart(ws1, (hist_start, hist_end), 4, 1,
                    "Tendencia Ingresos Netos", f"B{r}", GREEN_KPI)
    r += 14

    # ── SECCION: Egresos del Mes ──
    _style_section_header(ws1, r, 1, "Egresos del Mes", RED)
    r += 1
    _style_table_header(ws1, r, ["Categoria", "Monto", "Descripcion", "Fecha"])
    r += 1
    for i, eg in enumerate(egresos):
        is_alt = i % 2 == 1
        _style_data_row(ws1, r, [eg.categoria, float(eg.monto), eg.descripcion or "", str(eg.fecha)],
                        start_col=1, is_alt=is_alt, is_money_cols={2})
        r += 1
    # Total egresos
    _style_data_row(ws1, r, ["TOTAL EGRESOS", eg_mes_f,
                    "", ""], start_col=1, is_money_cols={2})
    for c in range(1, 5):
        ws1.cell(row=r, column=c).fill = PatternFill(
            start_color=RED, end_color=RED, fill_type="solid")
        ws1.cell(row=r, column=c).font = Font(bold=True, color=WHITE)
    r += 1
    # Ingreso Neto destacado
    r += 1
    ws1.merge_cells(f'A{r}:D{r}')
    neto_cell = ws1.cell(
        row=r, column=1, value=f"INGRESO NETO DEL MES: ${neto_mes:,.0f}")
    neto_cell.font = Font(bold=True, size=16,
                          color=GREEN if neto_mes >= 0 else RED)
    neto_cell.alignment = Alignment(horizontal='center')
    r += 2

    # ── SECCION: Detalle Ventas Bazar ──
    _style_section_header(ws1, r, 1, "Detalle de Ventas - Bazar", PURPLE_KPI)
    r += 1
    _style_table_header(
        ws1, r, ["ID", "Alumno", "Producto", "Cant.", "Total", "Estado", "Fecha"])
    r += 1
    for i, p in enumerate(pedidos):
        is_alt = i % 2 == 1
        estado_icon = "✅" if p.estado == "entregado" else (
            "⏳" if p.estado == "pendiente" else "❌")
        _style_data_row(ws1, r, [p.id, p.alumno_nombre, p.producto_nombre, p.cantidad, float(p.total), f"{estado_icon} {p.estado}", str(p.fecha_pedido)],
                        start_col=1, is_alt=is_alt, is_money_cols={5})
        r += 1
    # Total ventas
    _style_data_row(ws1, r, ["TOTAL VENTAS", "", "", "",
                    ventas_bazar, "", ""], start_col=1, is_money_cols={5})
    for c in range(1, 8):
        ws1.cell(row=r, column=c).fill = total_fill
        ws1.cell(row=r, column=c).font = total_font
    r += 2

    # ── SECCION: Comparativa Mes a Mes ──
    _style_section_header(ws1, r, 1, "Comparativa Mes a Mes", ORANGE_ACCENT)
    r += 1
    ws1.cell(row=r, column=1, value="Periodo 1:").font = Font(
        bold=True, size=11)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='right')
    # Data validation dropdown for month/year
    meses_opts = ",".join([f"'{h['mes']}'" for h in historico])
    dv_p1 = DataValidation(
        type="list", formula1=f'"{",".join([h["mes"] for h in historico])}"', allow_blank=True)
    dv_p1.error = "Selecciona un mes valido"
    dv_p1.errorTitle = "Mes invalido"
    ws1.add_data_validation(dv_p1)
    cell_p1 = ws1.cell(
        row=r, column=2, value=historico[-1]["mes"] if historico else "")
    dv_p1.add(cell_p1)
    r += 1
    ws1.cell(row=r, column=1, value="Periodo 2:").font = Font(
        bold=True, size=11)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='right')
    dv_p2 = DataValidation(
        type="list", formula1=f'"{",".join([h["mes"] for h in historico])}"', allow_blank=True)
    ws1.add_data_validation(dv_p2)
    cell_p2 = ws1.cell(
        row=r, column=2, value=historico[-2]["mes"] if len(historico) > 1 else "")
    dv_p2.add(cell_p2)
    r += 1

    # Tabla de comparativa
    r += 1
    tbl_headers_comp = ["Indicador", "Periodo 1", "Periodo 2", "Variacion"]
    _style_table_header(ws1, r, tbl_headers_comp)
    r += 1
    # Data con INDEX/MATCH formulas usando la hoja Historico Mensual (hidden)
    # Placeholder values - formulas reales referencian la hoja oculta
    indicadores = ["Ganancias (Neto)", "Ventas", "Egresos"]
    for i, ind in enumerate(indicadores):
        is_alt = i % 2 == 1
        _style_data_row(ws1, r, [ind, "", "", ""], start_col=1, is_alt=is_alt)
        r += 1

    # ── SECCION: Comparativa Ano a Ano ──
    r += 1
    _style_section_header(ws1, r, 1, "Comparativa Ano a Ano", TEAL_KPI)
    r += 1
    ws1.cell(row=r, column=1, value="Ano 1:").font = Font(bold=True, size=11)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='right')
    anos = list(set(h["anio"] for h in historico))
    anos.sort()
    dv_a1 = DataValidation(
        type="list", formula1=f'"{",".join(str(a) for a in anos)}"', allow_blank=True)
    ws1.add_data_validation(dv_a1)
    cell_a1 = ws1.cell(row=r, column=2, value=anos[-1] if anos else "")
    dv_a1.add(cell_a1)
    r += 1
    ws1.cell(row=r, column=1, value="Ano 2:").font = Font(bold=True, size=11)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='right')
    dv_a2 = DataValidation(
        type="list", formula1=f'"{",".join(str(a) for a in anos)}"', allow_blank=True)
    ws1.add_data_validation(dv_a2)
    cell_a2 = ws1.cell(
        row=r, column=2, value=anos[-2] if len(anos) >= 2 else "")
    dv_a2.add(cell_a2)
    r += 1
    r += 1
    _style_table_header(ws1, r, ["Indicador", "Ano 1", "Ano 2", "Variacion"])
    r += 1
    for i, ind in enumerate(indicadores):
        is_alt = i % 2 == 1
        _style_data_row(ws1, r, [ind, "", "", ""], start_col=1, is_alt=is_alt)
        r += 1

    # ================================================================
    # PESTANA 2: DETALLE PLANES
    # ================================================================
    ws2 = wb.create_sheet("Detalle Planes", 1)
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 22
    ws2.column_dimensions['E'].width = 22
    ws2.column_dimensions['F'].width = 22
    ws2.column_dimensions['G'].width = 22

    # Banner
    for c in range(1, 8):
        ws2.cell(row=1, column=c).fill = PatternFill(
            start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws2.merge_cells('A1:G1')
    ws2.cell(row=1, column=1, value="DETALLE DE PLANES").font = Font(
        bold=True, size=20, color=WHITE)
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 45
    for c in range(1, 8):
        ws2.cell(row=2, column=c).fill = PatternFill(
            start_color=ORANGE_ACCENT, end_color=ORANGE_ACCENT, fill_type="solid")
    ws2.row_dimensions[2].height = 6

    # Resumen por categoria
    r = 4
    _style_section_header(ws2, r, 1, "Resumen por Categoria", BLUE_KPI)
    r += 1
    _style_table_header(
        ws2, r, ["Categoria", "N° Planes", "Suscripciones Activas", "Ingresos del Mes"])
    r += 1
    for i, (key, cat) in enumerate(categorias.items()):
        is_alt = i % 2 == 1
        _style_data_row(ws2, r, [cat["label"], cat["planes"], cat["suscripciones"], cat["ingresos"]],
                        start_col=1, is_alt=is_alt, is_money_cols={4})
        r += 1
    r += 2

    # Listado completo de planes
    _style_section_header(ws2, r, 1, "Listado Completo de Planes", GREEN_KPI)
    r += 1
    _style_table_header(ws2, r, ["Nombre", "Genero", "Es Estudiante",
                        "Precio CLP", "Creditos", "Estado", "Susc. Activas"])
    r += 1
    for i, pl in enumerate(planes_list):
        is_alt = i % 2 == 1
        estado_txt = "Activo" if pl["activo"] else "Inactivo"
        ee_txt = "Si" if pl["es_estudiante"] else "No"
        _style_data_row(ws2, r, [pl["nombre"], pl["genero"], ee_txt, pl["precio"], pl["creditos"], estado_txt, pl["suscripciones"]],
                        start_col=1, is_alt=is_alt, is_money_cols={4})
        r += 1
    r += 2

    # Grafico de barras: suscripciones por categoria
    cat_data_start = r
    ws2.cell(row=r, column=1, value="Categoria").font = header_font
    ws2.cell(row=r, column=2, value="Suscripciones").font = header_font
    r += 1
    for i, (key, cat) in enumerate(categorias.items()):
        ws2.cell(row=r, column=1, value=cat["label"])
        ws2.cell(row=r, column=2, value=cat["suscripciones"])
        r += 1
    _add_bar_chart(ws2, (cat_data_start + 1, r - 1), 2, 1,
                   "Suscripciones Activas por Categoria", f"D{cat_data_start}", BLUE_KPI)

    # ================================================================
    # PESTANA 3: BAZAR Y SERVICIOS
    # ================================================================
    ws3 = wb.create_sheet("Bazar y Servicios", 2)
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 3
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 22
    ws3.column_dimensions['D'].width = 22
    ws3.column_dimensions['E'].width = 22
    ws3.column_dimensions['F'].width = 22
    ws3.column_dimensions['G'].width = 22

    # Banner
    for c in range(1, 8):
        ws3.cell(row=1, column=c).fill = PatternFill(
            start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws3.merge_cells('A1:G1')
    ws3.cell(row=1, column=1, value="BAZAR Y SERVICIOS").font = Font(
        bold=True, size=20, color=WHITE)
    ws3.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 45
    for c in range(1, 8):
        ws3.cell(row=2, column=c).fill = PatternFill(
            start_color=ORANGE_ACCENT, end_color=ORANGE_ACCENT, fill_type="solid")
    ws3.row_dimensions[2].height = 6

    # KPI cards
    r = 4
    _apply_kpi_card(ws3, r, 2, "Ventas del Mes",
                    f"${ventas_bazar:,.0f}", "Total en pedidos", ORANGE_ACCENT, "🛒")
    _apply_kpi_card(ws3, r, 3, "Pedidos",
                    f"{len(pedidos)}", "Total del mes", BLUE_KPI, "📦")
    _apply_kpi_card(ws3, r, 4, "Completados",
                    f"{pedidos_completados}", "Entregados", GREEN_KPI, "✅")
    _apply_kpi_card(ws3, r, 5, "Pendientes",
                    f"{pedidos_pendientes}", "Sin entregar", PURPLE_KPI, "⏳")
    r += 4

    # Detalle de pedidos
    _style_section_header(ws3, r, 1, "Detalle de Pedidos", BLUE_KPI)
    r += 1
    _style_table_header(
        ws3, r, ["ID", "Alumno", "Producto", "Cant.", "Total", "Estado", "Fecha"])
    r += 1
    for i, p in enumerate(pedidos):
        is_alt = i % 2 == 1
        estado_icon = "✅" if p.estado == "entregado" else (
            "⏳" if p.estado == "pendiente" else "❌")
        _style_data_row(ws3, r, [p.id, p.alumno_nombre, p.producto_nombre, p.cantidad, float(p.total), f"{estado_icon} {p.estado}", str(p.fecha_pedido)],
                        start_col=1, is_alt=is_alt, is_money_cols={5})
        r += 1
    # Total
    _style_data_row(ws3, r, ["TOTAL", "", "", "", ventas_bazar,
                    "", ""], start_col=1, is_money_cols={5})
    for c in range(1, 8):
        ws3.cell(row=r, column=c).fill = total_fill
        ws3.cell(row=r, column=c).font = total_font
    r += 2

    # Grafico torta: ventas por producto
    if ventas_producto:
        _style_section_header(ws3, r, 1, "Ventas por Producto", GREEN_KPI)
        r += 1
        prod_start = r
        for i, (prod_nom, prod_total) in enumerate(ventas_producto.items()):
            ws3.cell(row=r, column=1, value=prod_nom)
            ws3.cell(row=r, column=2, value=prod_total)
            r += 1
        _add_pie_chart(ws3, (prod_start, r - 1), 2, 1,
                       "Distribucion Ventas por Producto", f"D{prod_start}")
        r += 14

    # Servicios Especiales/Eventos (placeholder)
    r += 1
    _style_section_header(
        ws3, r, 1, "Servicios Especiales / Eventos", PURPLE_KPI)
    r += 1
    ws3.merge_cells(f'A{r}:G{r}')
    ws3.cell(row=r, column=1, value="Nota: La funcionalidad de Servicios Especiales y Eventos aun no esta disponible.").font = Font(
        italic=True, size=11, color="9CA3AF")
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal='center')

    # ================================================================
    # PESTANA 4: FLUJO DE CAJA
    # ================================================================
    ws4 = wb.create_sheet("Flujo de Caja", 3)
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 3
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 22
    ws4.column_dimensions['D'].width = 22
    ws4.column_dimensions['E'].width = 22
    ws4.column_dimensions['F'].width = 22

    # Banner
    for c in range(1, 7):
        ws4.cell(row=1, column=c).fill = PatternFill(
            start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws4.merge_cells('A1:F1')
    ws4.cell(row=1, column=1, value="FLUJO DE CAJA").font = Font(
        bold=True, size=20, color=WHITE)
    ws4.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws4.row_dimensions[1].height = 45
    for c in range(1, 7):
        ws4.cell(row=2, column=c).fill = PatternFill(
            start_color=ORANGE_ACCENT, end_color=ORANGE_ACCENT, fill_type="solid")
    ws4.row_dimensions[2].height = 6

    # Tabla mensual
    r = 4
    _style_section_header(ws4, r, 1, "Flujo Mensual", TEAL_KPI)
    r += 1
    _style_table_header(
        ws4, r, ["Mes", "Entradas", "Salidas", "Flujo Neto", "Saldo Acumulado"])
    r += 1
    flujo_start = r
    saldo_acum = 0
    for i, h in enumerate(historico):
        is_alt = i % 2 == 1
        flujo_neto = h["ingresos"] - h["egresos"]
        saldo_acum += flujo_neto
        _style_data_row(ws4, r, [h["mes"], h["ingresos"], h["egresos"], flujo_neto, saldo_acum],
                        start_col=1, is_alt=is_alt, is_money_cols={2, 3, 4, 5})
        r += 1
    flujo_end = r - 1
    # Saldo final
    _style_data_row(ws4, r, ["SALDO ACTUAL", sum(h["ingresos"] for h in historico), sum(h["egresos"] for h in historico),
                             sum(h["ingresos"] - h["egresos"] for h in historico), saldo_acum],
                    start_col=1, is_money_cols={2, 3, 4, 5})
    for c in range(1, 6):
        ws4.cell(row=r, column=c).fill = total_fill
        ws4.cell(row=r, column=c).font = total_font
    r += 2

    # Grafico linea: Saldo Acumulado
    _add_line_chart(ws4, (flujo_start, flujo_end), 5, 1,
                    "Saldo Acumulado", f"B{r}", TEAL_KPI)
    r += 14

    # Grafico barras: Entradas vs Salidas
    # Preparamos datos auxiliares
    chart_r = r
    _style_section_header(
        ws4, chart_r, 1, "Entradas vs Salidas", ORANGE_ACCENT)
    chart_r += 1
    ws4.cell(row=chart_r, column=1, value="Mes").font = header_font
    ws4.cell(row=chart_r, column=2, value="Entradas").font = header_font
    ws4.cell(row=chart_r, column=3, value="Salidas").font = header_font
    chart_r += 1
    bar_start = chart_r
    for i, h in enumerate(historico):
        ws4.cell(row=chart_r, column=1, value=h["mes"])
        ws4.cell(row=chart_r, column=2, value=h["ingresos"])
        ws4.cell(row=chart_r, column=3, value=h["egresos"])
        chart_r += 1
    bar_end = chart_r - 1

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Entradas vs Salidas"
    bar_chart.style = 10
    bar_chart.width = 13
    bar_chart.height = 8
    data_ref = Reference(
        ws4, min_col=2, min_row=bar_start - 1, max_row=bar_end)
    cats_ref = Reference(ws4, min_col=1, min_row=bar_start, max_row=bar_end)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    if bar_chart.series:
        bar_chart.series[0].graphicalProperties.solidFill = GREEN_KPI
    if len(bar_chart.series) > 1:
        bar_chart.series[1].graphicalProperties.solidFill = RED
    ws4.add_chart(bar_chart, f"B{bar_end + 2}")

    # ── HOJA OCULTA: Historico Mensual (para formulas INDEX/MATCH) ──
    ws_hist = wb.create_sheet("Historico Mensual", 4)
    ws_hist.sheet_state = 'hidden'
    ws_hist_headers = ["Año", "Mes", "MesNum", "Ingresos Totales", "Alumnos Activos",
                       "Nuevos Alumnos", "MRR", "Ventas Totales", "Egresos"]
    ws_hist.column_dimensions['A'].width = 10
    ws_hist.column_dimensions['B'].width = 14
    ws_hist.column_dimensions['C'].width = 10
    ws_hist.column_dimensions['D'].width = 16
    ws_hist.column_dimensions['E'].width = 16
    ws_hist.column_dimensions['F'].width = 16
    ws_hist.column_dimensions['G'].width = 14
    ws_hist.column_dimensions['H'].width = 14
    ws_hist.column_dimensions['I'].width = 14

    for i, hdr in enumerate(ws_hist_headers):
        ws_hist.cell(row=1, column=i + 1, value=hdr).font = Font(bold=True)
    for i, h in enumerate(historico):
        ws_hist.cell(row=i + 2, column=1, value=h["anio"])
        ws_hist.cell(row=i + 2, column=2, value=h["mes"])
        ws_hist.cell(row=i + 2, column=3, value=h["mes_num"])
        ws_hist.cell(row=i + 2, column=4, value=h["ingresos"])
        ws_hist.cell(row=i + 2, column=5, value=h["alumnos"])
        ws_hist.cell(row=i + 2, column=6, value=h["nuevos"])
        ws_hist.cell(row=i + 2, column=7, value=h["mrr"])
        ws_hist.cell(row=i + 2, column=8, value=h["ventas"])
        ws_hist.cell(row=i + 2, column=9, value=h["egresos"])

    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
